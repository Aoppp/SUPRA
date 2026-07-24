"""Import data from 数据-草稿.xlsx into the SUPRA database.

Excel format: 56 columns, 3-row merged header, data starts row 4.
Drops all existing data and re-imports fresh.

Usage:
    python import_excel.py [path/to/数据-草稿.xlsx]
"""
import os
import re
import sys
import zipfile
from xml.etree import ElementTree as ET

import openpyxl

EXCEL_PATH = sys.argv[1] if len(sys.argv) > 1 else os.path.join(
    os.path.dirname(__file__), "data", "数据-草稿.xlsx"
)
IMAGES_DIR = os.path.join(os.path.dirname(__file__), "data", "images")

# Column mapping (1-indexed) for 数据-草稿.xlsx
COL = {
    "compound_type":      2,   # 化合物类型
    "compound_name":      3,   # 化合物
    "english_name":       4,   # 英文名称
    "compound_image":     5,   # 化合物图片 (WPS DISPIMG formula)
    "molecular_formula":  6,   # 分子式
    "molecular_weight":   7,   # 分子量（g/mol）
    "smiles":             8,   # SMILES号
    "cas":                9,   # CAS号
    # 10: 水溶性  11: logP  12: 生物利用度  13: 天然来源  14: 应用名称
    "is_cosmetic":        15,
    "cosmetic_note":      16,
    "is_drug":            17,
    "drug_note":          18,
    "is_food":            19,
    "food_note":          20,
    "food_category":      21,
    "food_daily_intake":  22,
    "regulations":        23,
    # 24: 单组份自组装  25: 组装参与组分数  26: 具体组装成分
    "component_count":    25,
    "assembly_drive_method": 27,  # 组装驱动方式
    "assembly_type":      28,   # 最终组装体类型
    "responsiveness":     29,
    "surface_modification": 30,
    "driving_force":      31,   # 组装驱动力分析
    "morphology":         32,   # 形貌类型
    "particle_size":      33,
    "size_note":          34,
    "size_source":        35,
    "aqueous_phase":      36,
    "organic_phase":      37,
    "solute":             38,
    "concentration":      39,
    "component_ratio":    40,
    "temperature":        41,
    "temperature_note":   42,
    "ph":                 43,
    "ph_note":            44,
    "stirring":           45,
    "assembly_time":      46,
    "biological_activity": 47,
    "preparation_method": 48,
    "molecular_char":     49,
    "notes":              50,
    "doi":                51,
}

DATA_START = 4


def split_items(text: str) -> list[str]:
    if not text:
        return []
    text = text.replace("；", ";").replace("\n", ";").replace("\r", ";")
    parts = [p.strip() for p in text.split(";") if p.strip()]
    cleaned = []
    for p in parts:
        p = re.sub(r'^\d+[\.\、\s]+', '', p).strip()
        if p and not p.startswith("—"):
            cleaned.append(p)
    return cleaned


def parse_size_range(size_text: str) -> tuple[float | None, float | None]:
    if not size_text:
        return None, None
    m = re.search(r'(\d+\.?\d*)\s*±\s*(\d+\.?\d*)', size_text)
    if m:
        return float(m.group(1)) - float(m.group(2)), float(m.group(1)) + float(m.group(2))
    m = re.search(r'(\d+\.?\d*)\s*[–\-~]\s*(\d+\.?\d*)', size_text)
    if m:
        return float(m.group(1)), float(m.group(2))
    m = re.search(r'(\d+\.?\d*)', size_text)
    if m:
        v = float(m.group(1))
        return v, v
    return None, None


def clean_cas(cas: str) -> str:
    cas = cas.replace("‑", "-").strip()
    m = re.search(r'(\d{2,7}-\d{2}-\d)', cas)
    if m:
        return m.group(1)
    return cas if cas and cas != "无" else ""


def build_wps_image_map(excel_path: str) -> dict[str, tuple[bytes, str]]:
    """Parse WPS DISPIMG image map: ID_xxx → (bytes, ext)."""
    id_to_rid: dict[str, str] = {}
    rid_to_file: dict[str, str] = {}

    with zipfile.ZipFile(excel_path, "r") as z:
        names = z.namelist()

        if "xl/cellimages.xml" in names:
            ns_xdr = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
            ns_r = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
            tree = ET.fromstring(z.read("xl/cellimages.xml"))
            for pic in tree.iter(f"{{{ns_xdr}}}pic"):
                cnvpr = pic.find(f"{{{ns_xdr}}}nvPicPr/{{{ns_xdr}}}cNvPr")
                ns_a = "http://schemas.openxmlformats.org/drawingml/2006/main"
                blip = pic.find(f"{{{ns_xdr}}}blipFill/{{{ns_a}}}blip")
                if cnvpr is not None and blip is not None:
                    img_id = cnvpr.get("name", "")
                    rid = blip.get(f"{{{ns_r}}}embed", "")
                    if img_id and rid:
                        id_to_rid[img_id] = rid

        rels_path = "xl/_rels/cellimages.xml.rels"
        if rels_path in names:
            tree = ET.fromstring(z.read(rels_path))
            for rel in tree:
                rid = rel.get("Id", "")
                target = rel.get("Target", "")
                if rid and target:
                    rid_to_file[rid] = target

        result: dict[str, tuple[bytes, str]] = {}
        for img_id, rid in id_to_rid.items():
            media_rel = rid_to_file.get(rid, "")
            if not media_rel:
                continue
            media_path = f"xl/{media_rel}"
            if media_path in names:
                img_bytes = z.read(media_path)
                ext = media_rel.rsplit(".", 1)[-1].lower()
                if ext in ("jpeg", "jpg"):
                    ext = "jpg"
                elif ext == "gif":
                    ext = "gif"
                else:
                    ext = "png"
                result[img_id] = (img_bytes, ext)

    print(f"WPS image map: {len(result)} images")
    return result


def extract_dispimg_id(cell_value) -> str | None:
    if not cell_value:
        return None
    m = re.search(r'DISPIMG\("(ID_[A-F0-9]+)"', str(cell_value), re.IGNORECASE)
    return m.group(1) if m else None


def import_data():
    from database import engine, Base, get_db
    from models import Assembly, BuildingBlock, DrivingForce, Morphology, Property, AssemblyDriveMethod

    if not os.path.exists(EXCEL_PATH):
        print(f"Excel not found: {EXCEL_PATH}")
        sys.exit(1)

    print(f"Loading {EXCEL_PATH}...")
    image_map = build_wps_image_map(EXCEL_PATH)

    # Drop and recreate all tables
    Base.metadata.drop_all(bind=engine)
    Base.metadata.create_all(bind=engine)
    db = next(get_db())

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=False)
    ws = wb["自组装数据收集表"]
    os.makedirs(IMAGES_DIR, exist_ok=True)

    # Track images already saved for this CAS to reuse across rows
    cas_image_cache: dict[str, str] = {}

    imported = 0
    for row_idx in range(DATA_START, ws.max_row + 1):
        def cell(c):
            v = ws.cell(row=row_idx, column=c).value
            return str(v).strip() if v is not None else ""

        compound_name = cell(COL["compound_name"])
        if not compound_name:
            break

        # -- Compound type as building block --
        ct = cell(COL["compound_type"])
        bb = None
        if ct:
            from models import BuildingBlock as BB
            bb = db.query(BB).filter_by(name=ct).first()
            if not bb:
                bb = BB(name=ct)
                db.add(bb)
                db.flush()

        # -- Morphology --
        morph_raw = cell(COL["morphology"]).split("\n")[0].split("；")[0].split(";")[0].strip()
        morph_name = re.sub(r'^\d+[\.\、\s]+', '', morph_raw).strip()
        morph = None
        if morph_name:
            morph = db.query(Morphology).filter_by(name=morph_name).first()
            if not morph:
                morph = Morphology(name=morph_name, description=morph_name)
                db.add(morph)
                db.flush()

        # -- Assembly drive method --
        dm_name = cell(COL["assembly_drive_method"])
        drive_method = None
        if dm_name:
            drive_method = db.query(AssemblyDriveMethod).filter_by(name=dm_name).first()
            if not drive_method:
                drive_method = AssemblyDriveMethod(name=dm_name)
                db.add(drive_method)
                db.flush()

        # -- Driving forces --
        driving_forces = []
        for df_name in split_items(cell(COL["driving_force"])):
            df = db.query(DrivingForce).filter_by(name=df_name).first()
            if not df:
                df = DrivingForce(name=df_name)
                db.add(df)
                db.flush()
            driving_forces.append(df)

        # -- Properties (from biological activity) --
        bio_text = cell(COL["biological_activity"])
        properties = []
        for p_name in split_items(bio_text):
            if len(p_name) > 200:
                p_name = p_name[:200]
            p = db.query(Property).filter_by(name=p_name).first()
            if not p:
                p = Property(name=p_name)
                db.add(p)
                db.flush()
            properties.append(p)

        # -- Size --
        size_text = cell(COL["particle_size"])
        size_min, size_max = parse_size_range(size_text)

        # -- CAS --
        cas = clean_cas(cell(COL["cas"]))

        # -- Molecular weight --
        mw_raw = ws.cell(row=row_idx, column=COL["molecular_weight"]).value
        mol_weight = None
        if mw_raw is not None:
            try:
                mol_weight = float(mw_raw)
            except (TypeError, ValueError):
                pass

        def opt(k, default=None):
            v = cell(k) if isinstance(k, int) else cell(COL[k])
            if not v or v == "无":
                return default
            return v

        # -- Extract compound image (WPS DISPIMG) --
        compound_image = None
        img_cell_val = ws.cell(row=row_idx, column=COL["compound_image"]).value
        img_id = extract_dispimg_id(img_cell_val)

        if cas and cas in cas_image_cache:
            compound_image = cas_image_cache[cas]
        elif img_id and img_id in image_map:
            img_bytes, ext = image_map[img_id]
            safe = re.sub(r"[^\w\-]", "_", cas or compound_name)
            img_filename = f"{safe}.{ext}"
            img_path = os.path.join(IMAGES_DIR, img_filename)
            with open(img_path, "wb") as f:
                f.write(img_bytes)
            compound_image = f"/images/{img_filename}"
            if cas:
                cas_image_cache[cas] = compound_image

        assembly = Assembly(
            name=compound_name,
            english_name=opt(COL["english_name"]),
            compound_image=compound_image,
            compound_type=ct or None,
            molecular_weight=mol_weight,
            smiles=opt(COL["smiles"]),
            cas_number=cas or None,
            assembly_type=opt(COL["assembly_type"]),
            particle_size=size_text or None,
            aqueous_phase=opt(COL["aqueous_phase"]),
            organic_phase=opt(COL["organic_phase"]),
            solute=opt(COL["solute"]),
            concentration=opt(COL["concentration"]),
            component_ratio=opt(COL["component_ratio"]),
            preparation_method=opt(COL["preparation_method"]),
            size_nm_min=size_min,
            size_nm_max=size_max,
            size_note=opt(COL["size_note"]),
            size_source=opt(COL["size_source"]),
            doi=opt(COL["doi"]),
            biological_activity=bio_text or None,
            assembly_temperature=opt(COL["temperature"]),
            temperature_note=opt(COL["temperature_note"]),
            ph_value=opt(COL["ph"]),
            ph_note=opt(COL["ph_note"]),
            stirring_condition=opt(COL["stirring"]),
            assembly_time=opt(COL["assembly_time"]),
            molecular_characteristics=opt(COL["molecular_char"]),
            notes=opt(COL["notes"]),
            is_cosmetic=cell(COL["is_cosmetic"]) == "是",
            cosmetic_note=opt(COL["cosmetic_note"]),
            is_drug=cell(COL["is_drug"]) == "是",
            drug_note=opt(COL["drug_note"]),
            is_food=cell(COL["is_food"]) == "是",
            food_note=opt(COL["food_note"]),
            food_category=opt(COL["food_category"]),
            food_daily_intake=opt(COL["food_daily_intake"]),
            regulations=opt(COL["regulations"]),
            component_count=opt(COL["component_count"]),
            responsiveness=opt(COL["responsiveness"]),
            surface_modification=opt(COL["surface_modification"]),
            building_block_id=bb.id if bb else None,
            morphology_id=morph.id if morph else None,
            assembly_drive_method_id=drive_method.id if drive_method else None,
        )
        assembly.driving_forces = driving_forces
        assembly.properties = properties
        db.add(assembly)

        imported += 1
        if imported % 20 == 0:
            db.flush()
            print(f"  Imported {imported} rows...")

    db.commit()
    db.close()

    db2 = next(get_db())
    from models import Assembly as A
    print(f"\nDone! Imported {imported} assemblies.")
    print(f"  Distinct compounds: {db2.query(A.name).distinct().count()}")
    print(f"  Images saved: {len(cas_image_cache)}")
    db2.close()


if __name__ == "__main__":
    import_data()
