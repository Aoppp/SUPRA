"""Extract WPS-embedded images and supplemental fields from 数据-草稿.xlsx.

Updates existing DB records (matched by CAS number) with:
- compound_type (化合物类型)
- molecular_weight (分子量)
- compound_image (extracted from WPS DISPIMG cells)

Usage:
    python extract_images.py [path/to/数据-草稿.xlsx]
"""
import os
import re
import sys
import zipfile
from xml.etree import ElementTree as ET

import openpyxl

EXCEL_PATH = sys.argv[1] if len(sys.argv) > 1 else os.path.join(
    os.path.dirname(__file__), "..", "data", "数据-草稿.xlsx"
)
IMAGES_DIR = os.path.join(os.path.dirname(__file__), "data", "images")
os.makedirs(IMAGES_DIR, exist_ok=True)

# Column mapping for 数据-草稿.xlsx (1-indexed)
COL = {
    "compound_type": 2,
    "name": 3,
    "english_name": 4,
    "compound_image": 5,   # contains =_xlfn.DISPIMG("ID_xxx",1)
    "molecular_formula": 6,
    "molecular_weight": 7,
    "smiles": 8,
    "cas": 9,
    "is_cosmetic": 15,
    "is_drug": 17,
    "is_food": 19,
    "component_count": 25,
    "assembly_drive_method": 27,
    "assembly_type": 28,
    "responsiveness": 29,
    "surface_modification": 30,
    "driving_force": 31,
    "morphology": 32,
    "particle_size": 33,
    "size_note": 34,
    "size_source": 35,
    "aqueous_phase": 36,
    "organic_phase": 37,
    "solute": 38,
    "concentration": 39,
    "component_ratio": 40,
    "temperature": 41,
    "temperature_note": 42,
    "ph": 43,
    "ph_note": 44,
    "stirring": 45,
    "assembly_time": 46,
    "biological_activity": 47,
    "preparation_method": 48,
    "molecular_char": 49,
    "notes": 50,
    "doi": 51,
}

DATA_START_ROW = 4


def build_image_map(excel_path: str) -> dict[str, tuple[bytes, str]]:
    """Build mapping from WPS image ID (ID_xxx) to (image_bytes, extension)."""
    id_to_rid: dict[str, str] = {}
    rid_to_file: dict[str, str] = {}

    with zipfile.ZipFile(excel_path, "r") as z:
        names = z.namelist()

        # Parse xl/cellimages.xml: name="ID_xxx" r:embed="rIdN"
        if "xl/cellimages.xml" in names:
            ns = {
                "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
                "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
            }
            tree = ET.fromstring(z.read("xl/cellimages.xml"))
            for pic in tree.iter("{%s}pic" % ns["xdr"]):
                cnvpr = pic.find(
                    "{%s}nvPicPr/{%s}cNvPr" % (ns["xdr"], ns["xdr"])
                )
                blip = pic.find(
                    "{%s}blipFill/{http://schemas.openxmlformats.org/drawingml/2006/main}blip"
                    % ns["xdr"]
                )
                if cnvpr is not None and blip is not None:
                    img_id = cnvpr.get("name", "")  # e.g. "ID_EF26B4..."
                    rid = blip.get("{%s}embed" % ns["r"], "")
                    if img_id and rid:
                        id_to_rid[img_id] = rid

        # Parse xl/_rels/cellimages.xml.rels: rIdN → media/imageN.ext
        rels_path = "xl/_rels/cellimages.xml.rels"
        if rels_path in names:
            tree = ET.fromstring(z.read(rels_path))
            for rel in tree:
                rid = rel.get("Id", "")
                target = rel.get("Target", "")
                if rid and target:
                    rid_to_file[rid] = target  # e.g. "media/image1.png"

        # Read image bytes
        id_to_image: dict[str, tuple[bytes, str]] = {}
        for img_id, rid in id_to_rid.items():
            media_rel = rid_to_file.get(rid, "")
            if not media_rel:
                continue
            media_path = f"xl/{media_rel}"
            if media_path in names:
                img_bytes = z.read(media_path)
                ext = media_rel.rsplit(".", 1)[-1].lower()
                if ext == "gif":
                    ext = "gif"
                elif ext in ("jpeg", "jpg"):
                    ext = "jpg"
                else:
                    ext = "png"
                id_to_image[img_id] = (img_bytes, ext)

    print(f"Found {len(id_to_image)} WPS cell images")
    return id_to_image


def extract_dispimg_id(cell_value) -> str | None:
    """Extract ID_xxx from =_xlfn.DISPIMG("ID_xxx",1) formula string."""
    if not cell_value:
        return None
    s = str(cell_value)
    m = re.search(r'DISPIMG\("(ID_[A-F0-9]+)"', s, re.IGNORECASE)
    if m:
        return m.group(1)
    return None


def main():
    from database import get_db
    from models import Assembly

    if not os.path.exists(EXCEL_PATH):
        print(f"Excel not found: {EXCEL_PATH}")
        sys.exit(1)

    print(f"Loading {EXCEL_PATH}...")
    image_map = build_image_map(EXCEL_PATH)

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=False)
    ws = wb["自组装数据收集表"]

    db = next(get_db())
    updated = 0
    skipped = 0
    img_saved = 0

    for row_idx in range(DATA_START_ROW, ws.max_row + 1):
        def cell(c):
            v = ws.cell(row=row_idx, column=c).value
            return str(v).strip() if v is not None else ""

        name = cell(COL["name"])
        if not name:
            break

        cas_raw = cell(COL["cas"])
        # Normalize CAS
        cas_raw = cas_raw.replace("‑", "-").strip()
        m = re.search(r"(\d{2,7}-\d{2}-\d)", cas_raw)
        cas = m.group(1) if m else cas_raw

        # Find matching Assembly in DB
        assembly = None
        if cas:
            assembly = db.query(Assembly).filter(Assembly.cas_number == cas).first()
        if not assembly:
            assembly = db.query(Assembly).filter(Assembly.name == name).first()

        if not assembly:
            print(f"  Row {row_idx}: no match for '{name}' CAS={cas}, skipping")
            skipped += 1
            continue

        # compound_type
        ctype = cell(COL["compound_type"])
        if ctype:
            assembly.compound_type = ctype

        # molecular_weight
        mw_raw = ws.cell(row=row_idx, column=COL["molecular_weight"]).value
        if mw_raw is not None:
            try:
                assembly.molecular_weight = float(mw_raw)
            except (TypeError, ValueError):
                pass

        # compound_image via WPS DISPIMG
        img_cell_value = ws.cell(row=row_idx, column=COL["compound_image"]).value
        img_id = extract_dispimg_id(img_cell_value)
        if img_id and img_id in image_map:
            img_bytes, ext = image_map[img_id]
            safe_cas = re.sub(r"[^\w\-]", "_", cas or name)
            img_filename = f"{safe_cas}.{ext}"
            img_path = os.path.join(IMAGES_DIR, img_filename)
            with open(img_path, "wb") as f:
                f.write(img_bytes)
            assembly.compound_image = f"/images/{img_filename}"
            img_saved += 1

        db.add(assembly)
        updated += 1

    db.commit()
    db.close()
    print(f"\nDone. Updated {updated} records, {img_saved} images saved, {skipped} skipped.")


if __name__ == "__main__":
    main()
