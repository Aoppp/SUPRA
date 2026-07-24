from fastapi import FastAPI, Depends, Query, HTTPException, UploadFile, File, Form, Request
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy.orm import Session
from typing import Optional
import io
import os
import shutil
from datetime import datetime

import openpyxl

from database import engine, Base, get_db, SessionLocal

import crud
import schemas
import auth
from models import VisitLog

app = FastAPI(title="SUPRA API", description="Supramolecular Assembly Database API", version="0.2.0")

CORS_ORIGINS = os.getenv(
    "CORS_ORIGINS",
    "http://localhost:5173,http://localhost:3000",
).split(",")

app.add_middleware(
    CORSMiddleware,
    allow_origins=CORS_ORIGINS,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.middleware("http")
async def visit_log_middleware(request: Request, call_next):
    if request.url.path.startswith("/api/") and not request.url.path.startswith("/api/admin/"):
        db = SessionLocal()
        try:
            ip = request.headers.get("X-Forwarded-For", request.client.host if request.client else "unknown")
            if "," in ip:
                ip = ip.split(",")[0].strip()
            referer = request.headers.get("referer")
            ua = request.headers.get("user-agent")
            crud.log_visit(db, ip, request.url.path, ua, referer)
        except Exception:
            pass
        finally:
            db.close()
    response = await call_next(request)
    return response


@app.on_event("startup")
def startup():
    Base.metadata.create_all(bind=engine)
    # Migrate existing assemblies table: add view_count column if missing
    try:
        import sqlite3
        conn = sqlite3.connect(os.path.join(os.path.dirname(__file__), "data", "dolphin.db"))
        cols = [c[1] for c in conn.execute("PRAGMA table_info(assemblies)").fetchall()]
        if "view_count" not in cols:
            conn.execute("ALTER TABLE assemblies ADD COLUMN view_count INTEGER DEFAULT 0")
        if "compound_type" not in cols:
            conn.execute("ALTER TABLE assemblies ADD COLUMN compound_type VARCHAR(100)")
        if "molecular_weight" not in cols:
            conn.execute("ALTER TABLE assemblies ADD COLUMN molecular_weight FLOAT")
        conn.commit()
        conn.close()
    except Exception:
        pass  # not SQLite or DB doesn't exist yet


@app.get("/api/health")
def health():
    return {"status": "ok"}


@app.get("/api/compound-types", response_model=list[schemas.CompoundTypeCount])
def list_compound_types(db: Session = Depends(get_db)):
    return crud.get_compound_type_counts(db)


@app.get("/api/compounds", response_model=schemas.CompoundGroupResult)
def search_compounds(
    name: Optional[str] = Query(None),
    compound_type: Optional[str] = Query(None),
    assembly_type: Optional[str] = Query(None),
    is_cosmetic: Optional[bool] = Query(None),
    is_drug: Optional[bool] = Query(None),
    is_food: Optional[bool] = Query(None),
    size_min: Optional[float] = Query(None),
    size_max: Optional[float] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
):
    params = schemas.SearchParams(
        name=name, compound_type=compound_type,
        assembly_type=assembly_type,
        is_cosmetic=is_cosmetic, is_drug=is_drug, is_food=is_food,
        size_min=size_min, size_max=size_max,
        page=page, page_size=page_size,
    )
    total, results = crud.search_compounds_grouped(db, params)
    items = [schemas.CompoundGroupItem(**r) for r in results]
    return schemas.CompoundGroupResult(total=total, page=page, page_size=page_size, results=items)


@app.get("/api/compounds/{compound_name}/assemblies", response_model=list[schemas.AssemblyListItem])
def get_compound_assemblies(compound_name: str, db: Session = Depends(get_db)):
    assemblies = crud.get_compound_assemblies(db, compound_name)
    items = [schemas.AssemblyListItem.model_validate(a) for a in assemblies]
    for item, a in zip(items, assemblies):
        item.category = crud.compute_category(a)
        item.foodmate_url = crud.compute_foodmate_url(a)
    return items


@app.get("/api/search", response_model=schemas.SearchResult)
def search(
    name: Optional[str] = Query(None),
    compound_type: Optional[str] = Query(None),
    building_block: Optional[str] = Query(None),
    morphology: Optional[str] = Query(None),
    driving_force: Optional[str] = Query(None),
    property: Optional[str] = Query(None),
    assembly_type: Optional[str] = Query(None),
    assembly_drive_method: Optional[str] = Query(None),
    aqueous_phase: Optional[str] = Query(None),
    organic_phase: Optional[str] = Query(None),
    is_cosmetic: Optional[bool] = Query(None),
    is_drug: Optional[bool] = Query(None),
    is_food: Optional[bool] = Query(None),
    responsiveness: Optional[str] = Query(None),
    surface_modification: Optional[str] = Query(None),
    size_min: Optional[float] = Query(None),
    size_max: Optional[float] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
):
    params = schemas.SearchParams(
        name=name, compound_type=compound_type, building_block=building_block, morphology=morphology,
        driving_force=driving_force, property=property,
        assembly_type=assembly_type, assembly_drive_method=assembly_drive_method,
        aqueous_phase=aqueous_phase, organic_phase=organic_phase,
        is_cosmetic=is_cosmetic, is_drug=is_drug, is_food=is_food,
        responsiveness=responsiveness, surface_modification=surface_modification,
        size_min=size_min, size_max=size_max,
        page=page, page_size=page_size,
    )
    total, results = crud.search_assemblies(db, params)
    items = [schemas.AssemblyListItem.model_validate(r) for r in results]
    for item in items:
        a = next((r for r in results if r.id == item.id), None)
        if a:
            item.category = crud.compute_category(a)
            item.foodmate_url = crud.compute_foodmate_url(a)
    return schemas.SearchResult(
        total=total, page=page, page_size=page_size,
        results=items,
    )


@app.get("/api/assemblies/{assembly_id}", response_model=schemas.AssemblyDetail)
def get_assembly(assembly_id: int, db: Session = Depends(get_db)):
    a = crud.get_assembly_detail(db, assembly_id)
    if not a:
        raise HTTPException(status_code=404, detail="Assembly not found")
    crud.increment_view_count(db, assembly_id)
    result = schemas.AssemblyDetail.model_validate(a)
    result.category = crud.compute_category(a)
    result.foodmate_url = crud.compute_foodmate_url(a)
    return result


@app.get("/api/building-blocks", response_model=list[schemas.BuildingBlockOut])
def list_building_blocks(db: Session = Depends(get_db)):
    return crud.get_building_block_list(db)


@app.get("/api/morphologies", response_model=list[schemas.MorphologyOut])
def list_morphologies(db: Session = Depends(get_db)):
    return crud.get_morphology_list(db)


@app.get("/api/driving-forces", response_model=list[schemas.DrivingForceOut])
def list_driving_forces(db: Session = Depends(get_db)):
    return crud.get_driving_force_list(db)


@app.get("/api/properties", response_model=list[schemas.PropertyOut])
def list_properties(db: Session = Depends(get_db)):
    return crud.get_property_list(db)


@app.get("/api/assembly-drive-methods", response_model=list[schemas.AssemblyDriveMethodOut])
def list_assembly_drive_methods(db: Session = Depends(get_db)):
    return crud.get_assembly_drive_method_list(db)


@app.post("/api/assemblies", response_model=schemas.AssemblyDetail)
def create_assembly(data: schemas.AssemblyCreate, db: Session = Depends(get_db)):
    return crud.create_assembly(db, data)


IMAGES_DIR = os.path.join(os.path.dirname(__file__), "data", "images")
os.makedirs(IMAGES_DIR, exist_ok=True)


@app.post("/api/upload-image")
async def upload_image(file: UploadFile = File(...)):
    """Upload a single image file, return the path."""
    if not file.filename:
        raise HTTPException(status_code=400, detail="No file selected")
    ext = os.path.splitext(file.filename)[1] or ".png"
    import uuid
    safe_name = f"{uuid.uuid4().hex}{ext}"
    dest = os.path.join(IMAGES_DIR, safe_name)
    with open(dest, "wb") as f:
        shutil.copyfileobj(file.file, f)
    return {"path": f"/images/{safe_name}"}


@app.post("/api/assemblies/batch")
async def batch_create_assemblies(file: UploadFile = File(...), db: Session = Depends(get_db)):
    if not file.filename or not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Please upload an Excel file (.xlsx or .xls)")

    contents = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents))
    except Exception:
        raise HTTPException(status_code=400, detail="Failed to parse Excel file")

    ws = wb.active
    if ws.max_row < 4:
        raise HTTPException(status_code=400, detail="Excel file has no data rows")

    # --- Extract embedded images ---
    row_images: dict[int, list[tuple[bytes, str]]] = {}

    # Primary: use openpyxl _images
    for img in ws._images:
        img_data = img._data()
        img_fmt = img.format or "png"
        af = img.anchor._from
        at = img.anchor.to if hasattr(img.anchor, "to") else None
        from_excel = af.row + 1
        to_excel = (at.row + 1) if at else from_excel
        for r in range(from_excel, to_excel + 1):
            row_images.setdefault(r, []).append((img_data, img_fmt))

    # Fallback: extract directly from ZIP (fixes Linux openpyxl bug)
    if not row_images:
        import zipfile
        from xml.etree import ElementTree as ET
        with zipfile.ZipFile(io.BytesIO(contents)) as z:
            drawing_rels = {}
            rels_path = "xl/drawings/_rels/drawing1.xml.rels"
            if rels_path in z.namelist():
                ns = {"r": "http://schemas.openxmlformats.org/package/2006/relationships"}
                tree = ET.fromstring(z.read(rels_path))
                for rel in tree:
                    rid = rel.get("Id")
                    target = rel.get("Target")
                    if target and "image" in target.lower():
                        drawing_rels[rid] = target.replace("../media/", "")

            drawing_path = "xl/drawings/drawing1.xml"
            if drawing_path in z.namelist():
                ns = {
                    "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
                    "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
                }
                tree = ET.fromstring(z.read(drawing_path))
                for anchor in tree.iter("{%s}twoCellAnchor" % ns["xdr"]):
                    from_el = anchor.find("xdr:from", ns)
                    to_el = anchor.find("xdr:to", ns)
                    if from_el is None:
                        continue
                    from_row = int(from_el.find("xdr:row", ns).text)
                    to_row = (int(to_el.find("xdr:row", ns).text) + 1) if to_el is not None else from_row + 1
                    for r in range(from_row + 1, to_row + 1):
                        for blip in anchor.iter("{%s}blip" % ns["a"]):
                            embed = blip.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed")
                            if embed and embed in drawing_rels:
                                media_file = drawing_rels[embed]
                                img_bytes = z.read(f"xl/media/{media_file}")
                                fmt = media_file.rsplit(".", 1)[-1]
                                row_images.setdefault(r, []).append((img_bytes, fmt))

    # Build column mapping from 3-row header
    header_map: dict[int, str] = {}
    for c in range(1, ws.max_column + 1):
        val = ws.cell(row=3, column=c).value or ws.cell(row=2, column=c).value or ws.cell(row=1, column=c).value
        key = str(val).strip() if val else f"col_{c}"
        header_map[c] = key

    # Normalize Chinese headers to English field names
    CH_HEADER_MAP = {
        "化合物名称": "name",
        "化合物英文名": "english_name",
        "英文名": "english_name",
        "化合物图片": "compound_image",
        "分子结构图": "compound_image",
        "SMILES": "smiles",
        "CAS号": "cas_number",
        "CAS": "cas_number",
        "组装体类型": "assembly_type",
        "粒径": "particle_size",
        "水相": "aqueous_phase",
        "有机相": "organic_phase",
        "溶质": "solute",
        "浓度": "concentration",
        "组分比例": "component_ratio",
        "制备方法": "preparation_method",
        "最小粒径": "size_nm_min",
        "最大粒径": "size_nm_max",
        "粒径备注": "size_note",
        "粒径来源": "size_source",
        "DOI": "doi",
        "生物活性": "biological_activity",
        "组装温度": "assembly_temperature",
        "温度备注": "temperature_note",
        "pH值": "ph_value",
        "pH": "ph_value",
        "pH备注": "ph_note",
        "搅拌条件": "stirring_condition",
        "搅拌": "stirring_condition",
        "组装时间": "assembly_time",
        "分子特征": "molecular_characteristics",
        "分子特征参数": "molecular_characteristics",
        "备注": "notes",
        "化妆品": "is_cosmetic",
        "化妆品备注": "cosmetic_note",
        "药品": "is_drug",
        "药品备注": "drug_note",
        "食品": "is_food",
        "食品备注": "food_note",
        "食品类别": "food_category",
        "每日摄入量": "food_daily_intake",
        "法规": "regulations",
        "法规信息": "regulations",
        "组分数": "component_count",
        "响应性": "responsiveness",
        "表面修饰": "surface_modification",
        "外部链接": "url",
        "网址": "url",
        "化合物类型": "building_block",
        "构建基元": "building_block",
        "形貌": "morphology",
        "驱动方式": "assembly_drive_method",
        "组装驱动方式": "assembly_drive_method",
        "驱动力": "driving_forces",
        "性质": "properties",
        "化合物浓度": "concentration",
    }

    rows: list[dict] = []
    for r in range(4, ws.max_row + 1):
        row: dict[str, str] = {}
        for c in range(1, ws.max_column + 1):
            key = header_map.get(c, f"col_{c}")
            # Normalize Chinese headers
            key = CH_HEADER_MAP.get(key, key)
            val = ws.cell(row=r, column=c).value
            row[key] = str(val).strip() if val is not None else ""

        # Attach image if present for this row
        if r in row_images:
            imgs = row_images[r]
            img_bytes, fmt = imgs[0]
            safe_name = f"batch_{r}.{fmt}"
            dest = os.path.join(IMAGES_DIR, safe_name)
            with open(dest, "wb") as f:
                f.write(img_bytes)
            row["compound_image"] = f"/images/{safe_name}"

        if any(v for v in row.values()):
            rows.append(row)

    if not rows:
        raise HTTPException(status_code=400, detail="No data rows found in Excel file")

    created, errors = crud.batch_create_assemblies(db, rows)
    return {"created": created, "errors": errors, "total_rows": len(rows)}


@app.delete("/api/assemblies/{assembly_id}")
def delete_assembly(assembly_id: int, db: Session = Depends(get_db)):
    a = crud.get_assembly_detail(db, assembly_id)
    if not a:
        raise HTTPException(status_code=404, detail="Assembly not found")
    # Clean up image file if stored locally
    if a.compound_image and a.compound_image.startswith("/images/"):
        img_path = os.path.join(IMAGES_DIR, os.path.basename(a.compound_image))
        if os.path.exists(img_path):
            os.remove(img_path)
    db.delete(a)
    db.commit()
    return {"ok": True}


@app.get("/api/search-by-cas")
def search_by_cas(cas: str = Query(...), db: Session = Depends(get_db)):
    """Find assemblies by CAS number (exact or partial match)."""
    results = crud.search_by_cas(db, cas)
    items = [schemas.AssemblyListItem.model_validate(r) for r in results]
    for item in items:
        r = next((x for x in results if x.id == item.id), None)
        if r:
            item.category = crud.compute_category(r)
            item.foodmate_url = crud.compute_foodmate_url(r)
    return items


@app.get("/api/workbench", response_model=list[schemas.WorkProgressOut])
def list_work_progress(db: Session = Depends(get_db)):
    return crud.get_work_progress_list(db)


@app.post("/api/workbench", response_model=schemas.WorkProgressOut)
async def upload_work_progress(
    person_name: str = Form(...),
    file_name: str = Form(...),
    file_type: str = Form("other"),
    description: str = Form(""),
    file: UploadFile | None = File(None),
    db: Session = Depends(get_db),
):
    file_path = None
    if file and file.filename:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_name = f"{timestamp}_{file.filename}"
        dest = os.path.join(crud.UPLOAD_DIR, safe_name)
        with open(dest, "wb") as f:
            shutil.copyfileobj(file.file, f)
        file_path = f"/uploads/{safe_name}"

    data = schemas.WorkProgressCreate(
        person_name=person_name,
        file_name=file_name,
        file_type=file_type,
        description=description,
    )
    return crud.create_work_progress(db, data, file_path)


@app.delete("/api/workbench/{wp_id}")
def delete_work_progress(wp_id: int, db: Session = Depends(get_db)):
    ok = crud.delete_work_progress(db, wp_id)
    if not ok:
        raise HTTPException(status_code=404, detail="Work progress entry not found")
    return {"ok": True}


@app.get("/api/structure-image/{assembly_id}")
def structure_image(assembly_id: int, db: Session = Depends(get_db)):
    """Generate a 2D chemical structure image from SMILES using RDKit."""
    from fastapi.responses import Response
    a = crud.get_assembly_detail(db, assembly_id)
    if not a or not a.smiles:
        raise HTTPException(status_code=404, detail="No SMILES available")

    try:
        from rdkit import Chem
        from rdkit.Chem.Draw import rdMolDraw2D
        import io

        mol = Chem.MolFromSmiles(a.smiles)
        if mol is None:
            raise HTTPException(status_code=404, detail="Invalid SMILES")

        drawer = rdMolDraw2D.MolDraw2DCairo(300, 200)
        opts = drawer.drawOptions()
        opts.addStereoAnnotation = True
        drawer.DrawMolecule(mol)
        drawer.FinishDrawing()
        png_data = drawer.GetDrawingText()

        return Response(content=png_data, media_type="image/png")
    except ImportError:
        raise HTTPException(status_code=500, detail="RDKit not available")


# Serve uploaded files and images
@app.get("/uploads/{filepath:path}")
def serve_upload(filepath: str):
    path = os.path.join(crud.UPLOAD_DIR, filepath)
    if not os.path.exists(path):
        raise HTTPException(status_code=404)
    from fastapi.responses import FileResponse
    return FileResponse(path)


@app.get("/images/{filepath:path}")
def serve_image(filepath: str):
    path = os.path.join(IMAGES_DIR, filepath)
    if not os.path.exists(path):
        raise HTTPException(status_code=404)
    from fastapi.responses import FileResponse
    return FileResponse(path)


# ── Admin routes ──────────────────────────────────────────────

@app.post("/api/admin/login")
def admin_login(data: schemas.LoginRequest):
    if not auth.verify_password(data.password):
        raise HTTPException(status_code=403, detail="Wrong password")
    token = auth.create_token()
    return {"token": token}


@app.get("/api/admin/visits", dependencies=[Depends(auth.require_admin)])
def admin_visits(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    total, results = crud.get_visits(db, page, page_size, date_from, date_to)
    items = [schemas.VisitLogOut.model_validate(r) for r in results]
    return schemas.VisitListResult(
        total=total, page=page, page_size=page_size, results=items,
    )


@app.get("/api/admin/stats", dependencies=[Depends(auth.require_admin)])
def admin_stats(
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    return crud.get_admin_stats(db, date_from, date_to)


@app.get("/api/admin/trend", dependencies=[Depends(auth.require_admin)])
def admin_trend(
    date_from: str = Query(...),
    date_to: str = Query(...),
    db: Session = Depends(get_db),
):
    daily = crud.get_trend_data(db, date_from, date_to)
    return schemas.TrendData(daily=daily)


@app.get("/api/admin/top-molecules", dependencies=[Depends(auth.require_admin)])
def admin_top_molecules(n: int = Query(20, ge=1, le=100), db: Session = Depends(get_db)):
    results = crud.get_top_molecules(db, n)
    return [schemas.TopMolecule.model_validate(r) for r in results]


@app.get("/api/admin/export-visits", dependencies=[Depends(auth.require_admin)])
def export_visits(db: Session = Depends(get_db)):
    from fastapi.responses import PlainTextResponse
    csv_data = crud.export_visits_csv(db)
    return PlainTextResponse(
        content=csv_data,
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=visit_logs.csv"},
    )


@app.get("/api/admin/export-molecule-stats", dependencies=[Depends(auth.require_admin)])
def export_molecule_stats(db: Session = Depends(get_db)):
    from fastapi.responses import PlainTextResponse
    csv_data = crud.export_molecule_stats_csv(db)
    return PlainTextResponse(
        content=csv_data,
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=molecule_stats.csv"},
    )
